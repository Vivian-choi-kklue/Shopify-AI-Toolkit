#!/usr/bin/env python3
from __future__ import annotations

import json
import os
import re
import sys
import time
import urllib.error
import urllib.request
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
API_VERSION = os.getenv("SHOPIFY_API_VERSION", "2026-04")
TZ = ZoneInfo(os.getenv("REPORT_TIMEZONE", "Asia/Shanghai"))

HEADERS = [
    "Name",
    "Date",
    "Subtotal",
    "Discount Amount",
    "Lineitem quantity",
    "Collection",
    "Lineitem name",
    "Lineitem price",
    "Quantity Price",
    "Lineitem sku",
    "Categories",
    "Material",
    "Billing Name",
    "Employee",
    "Location",
    "Source",
    "地區",
    "地區",
]

ORDER_LOCATION_SOURCE_OVERRIDES = {
    "#KJ7232": {"location": "Harbour City", "source": "pos"},
}


class ShopifyAccessFallback(Exception):
    def __init__(self, feature: str):
        super().__init__(feature)
        self.feature = feature


def clean_number(value: float | int | str | None):
    amount = float(value or 0)
    rounded = round(amount, 2)
    if abs(rounded - round(rounded)) < 0.00001:
        return int(round(rounded))
    return rounded


def money_from_set(value: dict | None) -> float:
    try:
        return float(value["shopMoney"]["amount"])
    except Exception:
        return 0.0


def country_to_region(country: str | None) -> str:
    if not country:
        return ""
    mapping = {
        "Hong Kong": "HK",
        "China": "CN",
        "United States": "US",
        "Japan": "JP",
        "Singapore": "SG",
        "Macau": "MO",
        "Macao": "MO",
        "Taiwan": "TW",
        "United Arab Emirates": "AE",
        "India": "IN",
        "Australia": "AU",
        "Canada": "CA",
    }
    return mapping.get(country, country)


def region_for(order: dict) -> str:
    phones = [
        order.get("phone"),
        (order.get("billingAddress") or {}).get("phone"),
        (order.get("shippingAddress") or {}).get("phone"),
    ]
    for raw_phone in phones:
        raw_phone = raw_phone or ""
        digits = re.sub(r"\D", "", raw_phone)
        if not digits:
            continue
        if digits.startswith("852") or (len(digits) == 8 and not raw_phone.strip().startswith("+")):
            return "HK"
        if digits.startswith("86"):
            return "CN"
        if digits.startswith("65"):
            return "SG"
        if raw_phone.strip().startswith("+63") or digits.startswith("0063"):
            return "PH"
        if digits.startswith("853"):
            return "MO"
        if digits.startswith("886"):
            return "TW"
        if digits.startswith("971"):
            return "AE"
        if digits.startswith("91"):
            return "IN"
        if digits.startswith("61"):
            return "AU"
        if digits.startswith("81"):
            return "JP"
        if digits.startswith("1") and len(digits) >= 11:
            billing = order.get("billingAddress") or {}
            shipping = order.get("shippingAddress") or {}
            return billing.get("countryCodeV2") or shipping.get("countryCodeV2") or "US/CA"
        break
    billing = order.get("billingAddress") or {}
    shipping = order.get("shippingAddress") or {}
    return billing.get("countryCodeV2") or shipping.get("countryCodeV2") or "無"


def infer_collection(name: str) -> str:
    lower = name.lower()
    if "paper bag" in lower or "packaging" in lower:
        return "Packaging"
    if "crystal beaded" in lower:
        return "Crystal DIY"
    if lower.startswith("unlock essential"):
        return "Unlock Essential"
    if lower.startswith("unlock marks"):
        return "Unlock Marks"
    if lower.startswith("unlock amulet"):
        return "Unlock Amulet"
    if lower.startswith("unlock"):
        return "Unlock Classic"
    for key, value in [
        ("d muse", "D Muse"),
        ("smiley", "Smiley"),
        ("shine", "Shine"),
        ("moment", "Moment"),
        ("essential", "Essential"),
        ("lokki", "Lokki"),
        ("sol", "SOL"),
        ("teddy", "Teddy"),
        ("raindrop", "Raindrop"),
        ("rain drop", "Raindrop"),
        ("together", "Together"),
        ("illume", "Illume"),
        ("horseshoe", "Horseshoe"),
        ("garden", "Garden"),
    ]:
        if lower.startswith(key):
            return value
    for key, value in [("smiley", "Smiley"), ("moment", "Moment"), ("essential", "Essential"), ("lokki", "Lokki"), ("shine", "Shine")]:
        if key in lower:
            return value
    return "Custom"


def infer_category(name: str) -> str:
    lower = name.lower()
    if "paper bag" in lower or "packaging" in lower:
        return "Packaging"
    if "printer kit" in lower:
        return "Custom"
    if "crystal beaded" in lower:
        return "Crystal DIY"
    if "necklace" in lower or "choker" in lower:
        return "Necklace"
    if "pendant" in lower:
        return "Pendant"
    if "charm" in lower:
        return "Charm"
    if "bracelet" in lower or "bangle" in lower:
        return "Bracelet"
    if "earring" in lower or "earrings" in lower:
        return "Earring"
    if "ring" in lower:
        return "Ring"
    return "Custom"


def infer_material(sku: str, name: str) -> str:
    lower = name.lower()
    sku = sku or ""
    if "paper bag" in lower or "packaging" in lower:
        return "Packaging"
    if "printer kit" in lower:
        return "Custom"
    if "platinum" in lower:
        return "Platinum"
    if "18k" in lower:
        return "18K"
    if "14k" in lower:
        return "14K"
    if "9k" in lower:
        return "9K"
    if "silver" in lower or "vermeil" in lower or re.search(r"(^|-)[A-Z]*S", sku):
        return "S925"
    if re.search(r"(^|-)[A-Z]*K", sku) or sku.startswith(("UC", "MED", "DM-")):
        return "18K"
    if re.search(r"(^|-)[A-Z]*F", sku) or sku.startswith(("FUC", "FUCD", "FMED")):
        return "14K"
    if re.search(r"(^|-)[A-Z]*N", sku):
        return "9K"
    if re.search(r"(^|-)[A-Z]*P", sku):
        return "Platinum"
    return "Custom"


def canonical_employee(alias: str, location: str = "") -> str:
    cleaned = re.sub(r"[\s$，,。:：的]+", "", alias or "").lower()
    mapping = {
        "becky": "Tam Becky",
        "tambecky": "Tam Becky",
        "idy": "Idy Leung",
        "idyleung": "Idy Leung",
        "joey": "Joey Chung",
        "joeychung": "Joey Chung",
        "bonnie": "Bonnie Ma",
        "bonniema": "Bonnie Ma",
        "catherine": "Lau Catherine",
        "laucatherine": "Lau Catherine",
        "esso": "Esso Chan",
        "essochan": "Esso Chan",
        "corsa": "Corsa Ko",
        "corsako": "Corsa Ko",
        "shirley": "Shirley Chan",
        "shirleychan": "Shirley Chan",
        "karza": "Karza Wong",
        "karzawong": "Karza Wong",
        "kitty": "Kitty Chui",
        "kittychui": "Kitty Chui",
    }
    if cleaned in ("may", "maychan", "阿may"):
        return "May Wong" if "harbour" in (location or "").lower() else "May Chan"
    for key, value in mapping.items():
        if key and key in cleaned:
            return value
    return alias.strip() or "無"


def employee_override_from_notes(note: str, current_employee: str, location: str = ""):
    compact = re.sub(r"\s+", "", note or "")
    if not compact:
        return current_employee or "無", None
    lower_compact = compact.lower()
    split_allocations = []
    for amount, alias in re.findall(r"\$?([0-9]+(?:\.[0-9]+)?)歸屬([A-Za-z\u4e00-\u9fff]+)", compact):
        split_allocations.append({"employee": canonical_employee(alias, location), "amount": clean_number(amount)})
    for alias, amount in re.findall(r"([A-Za-z\u4e00-\u9fff]+)生意額為\$?([0-9]+(?:\.[0-9]+)?)", compact):
        split_allocations.append({"employee": canonical_employee(alias, location), "amount": clean_number(amount)})
    for match in re.finditer(r"([A-Za-z]+)\$([0-9]+(?:\.[0-9]+)?)", compact):
        alias, amount = match.groups()
        before = compact[max(0, match.start() - 8):match.start()].lower()
        if "歸屬" in before:
            continue
        if alias.lower() in {"corsa", "joey", "idy", "bonnie", "catherine", "esso", "may", "karza"}:
            split_allocations.append({"employee": canonical_employee(alias, location), "amount": clean_number(amount)})
    if split_allocations:
        employees = []
        for item in split_allocations:
            if item["employee"] not in employees:
                employees.append(item["employee"])
        return " / ".join(employees), {"type": "split", "allocations": split_allocations}

    override_patterns = [
        (r"(?:全單|整單|此單)歸.*?joey", "Joey Chung"),
        (r"(?:全單|整單|此單)歸.*?bonnie", "Bonnie Ma"),
        (r"(?:全單|整單|此單)歸.*?may", canonical_employee("may", location)),
        (r"(?:全單|整單|此單)歸.*?catherine", "Lau Catherine"),
        (r"(?:全單|整單|此單)歸.*?esso", "Esso Chan"),
        (r"(?:全單|整單|此單)歸.*?corsa", "Corsa Ko"),
        (r"(?:全單|整單|此單)歸.*?shirley", "Shirley Chan"),
        (r"(?:全單|整單|此單)歸.*?karza", "Karza Wong"),
        (r"(?:全單|整單)是.*?may", canonical_employee("may", location)),
        (r"全部bonnie", "Bonnie Ma"),
        (r"(?:這單|此單)是阿?may", canonical_employee("may", location)),
        (r"(?:這單|此單)是catherine", "Lau Catherine"),
        (r"(?:這單|此單)是esso", "Esso Chan"),
        (r"(?:這單|此單)是joey", "Joey Chung"),
        (r"(?:這單|此單)是bonnie", "Bonnie Ma"),
        (r"(?:這單|此單)是corsa", "Corsa Ko"),
        (r"(?:這單|此單)是shirley", "Shirley Chan"),
        (r"(?:這單|此單)是karza", "Karza Wong"),
    ]
    for pattern, employee in override_patterns:
        if re.search(pattern, lower_compact):
            return employee, {"type": "full_order", "employee": employee}
    return current_employee or "無", None


def normalized_source(order: dict) -> str:
    name = order["name"]
    if name in ORDER_LOCATION_SOURCE_OVERRIDES:
        return ORDER_LOCATION_SOURCE_OVERRIDES[name]["source"]
    location = ((order.get("retailLocation") or {}).get("name") or "").strip()
    if location.lower() == "9-17 wing kut street":
        return "web"
    return order.get("sourceName") or "無"


def normalized_location(order: dict, source: str) -> str:
    name = order["name"]
    if name in ORDER_LOCATION_SOURCE_OVERRIDES:
        return ORDER_LOCATION_SOURCE_OVERRIDES[name]["location"]
    location = ((order.get("retailLocation") or {}).get("name") or "").strip()
    if location.lower() == "9-17 wing kut street":
        return "web"
    if not location and source == "web":
        return "web"
    return location or "無"


def is_packaging_line(line: dict) -> bool:
    name = (line.get("name") or line.get("title") or "").lower()
    sku = (line.get("sku") or "").upper()
    return "paper bag" in name or sku.startswith("CPT")


@dataclass
class ShopifyClient:
    store_domain: str
    token: str

    @property
    def endpoint(self) -> str:
        return f"https://{self.store_domain}/admin/api/{API_VERSION}/graphql.json"

    def graphql(self, query: str, variables: dict) -> dict:
        body = json.dumps({"query": query, "variables": variables}).encode("utf-8")
        request = urllib.request.Request(
            self.endpoint,
            data=body,
            headers={
                "X-Shopify-Access-Token": self.token,
                "Content-Type": "application/json",
                "Accept": "application/json",
            },
            method="POST",
        )
        try:
            with urllib.request.urlopen(request, timeout=45) as response:
                payload = json.loads(response.read().decode("utf-8"))
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"Shopify API HTTP {exc.code}: {detail}") from exc
        errors = payload.get("errors") or []
        if errors:
            text = json.dumps(errors)
            if "staffMember" in text and "Access denied" in text:
                raise ShopifyAccessFallback("staffMember")
            if "returns" in text and "Access denied" in text:
                raise ShopifyAccessFallback("returns")
            raise RuntimeError(text)
        return payload["data"]


def build_query(include_staff: bool, include_returns: bool) -> str:
    staff = "staffMember { name }" if include_staff else ""
    returns = """
      returns(first: 10) {
        nodes {
          name
          status
          returnLineItems(first: 50) {
            nodes {
              ... on ReturnLineItem {
                quantity
                refundedQuantity
                fulfillmentLineItem { lineItem { id sku title name } }
              }
            }
          }
          exchangeLineItems(first: 50) {
            nodes { lineItem { id sku title name quantity currentQuantity } }
          }
        }
      }
    """ if include_returns else ""
    return f"""
query OrdersForMonth($first: Int!, $after: String, $query: String!) {{
  orders(first: $first, after: $after, query: $query, sortKey: CREATED_AT) {{
    nodes {{
      id
      name
      createdAt
      processedAt
      displayFinancialStatus
      cancelledAt
      returnStatus
      sourceName
      note
      phone
      currentSubtotalPriceSet {{ shopMoney {{ amount currencyCode }} }}
      currentTotalDiscountsSet {{ shopMoney {{ amount currencyCode }} }}
      totalRefundedSet {{ shopMoney {{ amount currencyCode }} }}
      retailLocation {{ name }}
      billingAddress {{ name phone countryCodeV2 country }}
      shippingAddress {{ name phone countryCodeV2 country }}
      {staff}
      lineItems(first: 100) {{
        nodes {{
          id
          title
          name
          sku
          quantity
          currentQuantity
          originalUnitPriceSet {{ shopMoney {{ amount currencyCode }} }}
          originalTotalSet {{ shopMoney {{ amount currencyCode }} }}
          discountedTotalSet(withCodeDiscounts: true) {{ shopMoney {{ amount currencyCode }} }}
          totalDiscountSet {{ shopMoney {{ amount currencyCode }} }}
        }}
      }}
      refunds(first: 10) {{
        id
        totalRefundedSet {{ shopMoney {{ amount currencyCode }} }}
        refundLineItems(first: 50) {{
          nodes {{
            quantity
            lineItem {{ id sku title name }}
            subtotalSet {{ shopMoney {{ amount currencyCode }} }}
          }}
        }}
      }}
      {returns}
    }}
    pageInfo {{ hasNextPage endCursor }}
  }}
}}
"""


def fetch_orders(client: ShopifyClient, query_filter: str):
    include_staff = os.getenv("SHOPIFY_FETCH_STAFF", "1") != "0"
    include_returns = os.getenv("SHOPIFY_FETCH_RETURNS", "1") != "0"
    unavailable = []
    while True:
        try:
            return _fetch_orders(client, query_filter, include_staff, include_returns), unavailable
        except ShopifyAccessFallback as exc:
            if exc.feature == "staffMember" and include_staff:
                include_staff = False
                unavailable.append("staffMember")
                continue
            if exc.feature == "returns" and include_returns:
                include_returns = False
                unavailable.append("returns")
                continue
            raise


def _fetch_orders(client: ShopifyClient, query_filter: str, include_staff: bool, include_returns: bool) -> list[dict]:
    query = build_query(include_staff, include_returns)
    after = None
    orders = []
    while True:
        data = client.graphql(query, {"first": 25, "after": after, "query": query_filter})
        conn = data["orders"]
        orders.extend(conn["nodes"])
        page = conn["pageInfo"]
        if not page["hasNextPage"]:
            break
        after = page["endCursor"]
        time.sleep(0.25)
    return orders


def create_workbook(rows: list[list]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Original Data"
    widths = [14, 12, 14, 16, 18, 18, 46, 16, 16, 22, 16, 14, 24, 18, 18, 12, 10, 10]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(1, col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    ws.row_dimensions[1].height = 24
    center_cols = {1, 2, 5, 6, 10, 11, 12, 14, 15, 16, 17, 18}
    money_cols = {3, 4, 8, 9}
    for row_idx, values in enumerate(rows, 2):
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row_idx, col_idx)
            cell.value = value
            cell.alignment = Alignment(
                horizontal="center" if col_idx in center_cols else "left",
                vertical="center",
                wrap_text=col_idx in {7, 13},
            )
            if col_idx in money_cols:
                cell.number_format = '#,##0.00'
        ws.row_dimensions[row_idx].height = 22
    ws.auto_filter.ref = f"A1:R{max(1, len(rows) + 1)}"
    return wb


def build_rows(orders: list[dict]):
    output_rows = []
    excluded = []
    corrected = []
    employee_overrides = []
    unresolved = []

    for order in sorted(orders, key=lambda item: item["createdAt"]):
        name = order["name"]
        current_subtotal = money_from_set(order.get("currentSubtotalPriceSet"))
        current_discount = money_from_set(order.get("currentTotalDiscountsSet"))
        refunded = money_from_set(order.get("totalRefundedSet"))
        cancelled = bool(order.get("cancelledAt"))
        line_items = (order.get("lineItems") or {}).get("nodes") or []
        active_lines = [line for line in line_items if int(line.get("currentQuantity") or 0) > 0]

        if cancelled or current_subtotal <= 0.005:
            excluded.append({"name": name, "reason": "cancelled or current subtotal is 0", "refunded": clean_number(refunded)})
            continue
        if active_lines and all(is_packaging_line(line) for line in active_lines):
            excluded.append({"name": name, "reason": "packaging-only order"})
            continue
        if not active_lines:
            excluded.append({"name": name, "reason": "no active line items"})
            continue

        removed_lines = [
            {
                "sku": line.get("sku"),
                "name": line.get("name"),
                "quantity": int(line.get("quantity") or 0) - int(line.get("currentQuantity") or 0),
            }
            for line in line_items
            if int(line.get("quantity") or 0) > int(line.get("currentQuantity") or 0)
        ]
        if removed_lines:
            corrected.append({
                "name": name,
                "reason": "line item currentQuantity lower than original quantity",
                "removed_lines": removed_lines,
                "current_subtotal": clean_number(current_subtotal),
                "current_discount": clean_number(current_discount),
            })

        source = normalized_source(order)
        location = normalized_location(order, source)
        billing = order.get("billingAddress") or {}
        shipping = order.get("shippingAddress") or {}
        billing_name = billing.get("name") or shipping.get("name") or "無"
        staff_name = ((order.get("staffMember") or {}).get("name") or "").strip()
        employee = staff_name or ("web" if source == "web" else "無")
        employee, override = employee_override_from_notes(order.get("note") or "", employee, location)
        if override:
            employee_overrides.append({"name": name, "employee": employee, "note": order.get("note"), "override": override})
        elif employee == "無" and source == "pos":
            unresolved.append({"name": name, "field": "Employee", "reason": "POS staff unavailable from Shopify API and no note override"})

        region = region_for(order)
        if region == "無":
            note = re.sub(r"\s+", "", order.get("note") or "")
            if not re.search(r"不留資料|沒留資料|不能資料|不肯留資料", note):
                unresolved.append({"name": name, "field": "地區", "reason": "phone/country unavailable"})

        order_date = datetime.fromisoformat(order["createdAt"].replace("Z", "+00:00")).astimezone(TZ).strftime("%Y-%m-%d")
        for idx, line in enumerate(active_lines):
            qty = int(line.get("currentQuantity") or 0)
            line_name = (line.get("name") or line.get("title") or "無").strip()
            sku = (line.get("sku") or "").strip()
            price = money_from_set(line.get("originalUnitPriceSet"))
            qty_price = qty * price
            inferred = (infer_collection(line_name), infer_category(line_name), infer_material(sku, line_name))
            collection, category, material = inferred
            output_rows.append([
                name,
                order_date,
                clean_number(current_subtotal) if idx == 0 else None,
                clean_number(current_discount) if idx == 0 else None,
                clean_number(qty),
                collection or inferred[0] or "Custom",
                line_name,
                clean_number(price),
                clean_number(qty_price),
                sku or "無",
                category or inferred[1] or "Custom",
                material or inferred[2] or "Custom",
                billing_name,
                employee,
                location,
                source,
                region,
                region,
            ])

    return output_rows, excluded, corrected, employee_overrides, unresolved


def qa(rows: list[list]) -> dict:
    required_indexes = [5, 8, 10, 11, 12, 13, 14, 15, 16, 17]
    blanks = []
    for row_num, row in enumerate(rows, 2):
        for idx in required_indexes:
            if row[idx] is None or str(row[idx]).strip() == "":
                blanks.append({"row": row_num, "order": row[0], "column": HEADERS[idx]})
    return {
        "row_count": len(rows),
        "order_count": len({row[0] for row in rows}),
        "subtotal_total": clean_number(sum(float(row[2] or 0) for row in rows)),
        "required_blank_count": len(blanks),
        "required_blank_samples": blanks[:20],
    }


def main() -> int:
    store_domain = os.getenv("SHOPIFY_STORE_DOMAIN", "").strip()
    token = os.getenv("SHOPIFY_ADMIN_TOKEN", "").strip()
    if not store_domain or not token:
        print("Missing SHOPIFY_STORE_DOMAIN or SHOPIFY_ADMIN_TOKEN.", file=sys.stderr)
        return 2

    now = datetime.now(TZ)
    start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    end = now + timedelta(days=1)
    query_filter = f"created_at:>={start.strftime('%Y-%m-%d')} created_at:<{end.strftime('%Y-%m-%d')}"
    client = ShopifyClient(store_domain=store_domain, token=token)
    orders, unavailable_features = fetch_orders(client, query_filter)
    rows, excluded, corrected, employee_overrides, unresolved = build_rows(orders)

    year_dir = DATA_DIR / str(now.year)
    year_dir.mkdir(parents=True, exist_ok=True)
    month_name = now.strftime("%B").upper()
    workbook_path = year_dir / f"{now.year} {month_name} Sales Data.xlsx"
    summary_path = year_dir / f"summary-{now.strftime('%Y-%m-%d')}.json"
    workbook = create_workbook(rows)
    workbook.save(workbook_path)

    summary = {
        "run_at": now.isoformat(),
        "query_filter": query_filter,
        "shopify_api_version": API_VERSION,
        "unavailable_features": unavailable_features,
        "output_workbook": str(workbook_path.relative_to(ROOT)),
        "qa": qa(rows),
        "excluded_orders": excluded,
        "corrected_return_exchange_orders": corrected,
        "employee_note_overrides": employee_overrides,
        "manual_review": unresolved,
    }
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
